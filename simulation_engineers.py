'''
Author: Alejandro Bautista Ramos

The following is a python file that tries to simulate the behaviour of each engineer from the
COL_TriageTickets_ExportV2_01_04_2018_modified_data_analysis.xlsx.
'''
from numpy.random import choice
import numpy as np
import random
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from datetime import date, timedelta, datetime
from random import randrange


class Engineer:

    def __init__(self,
                 first_name,
                 last_name,
                 amount_of_tickets_solved_day,
                 elements_category = None,
                 weight_category   = None,
                 elements_problem_category = None,
                 weight_problem_category = None,
                 elements_affected_devices = None,
                 weight_affected_devices = None,
                 elements_reason_for_creating = None,
                 weights_reason_for_creating  = None,
                 elements_issue_status = None,
                 weights_issue_status = None,
                 elements_sla_met = None,
                 weights_sla_met  = None,
                 elements_priority = None,
                 weights_priority = None,
                 #elements_kpi_i = None,
                 #weights_kpi_i = None
                 ):

        ''' Define the Engineer constructor. If one variable is non-public then it must contain the getters and setters
            properties. '''

        self.first_name                   = first_name
        self.last_name                    = last_name
        self.amount_of_tickets_solved_day = amount_of_tickets_solved_day # non-public attribute, can be changed because there is a setter

        if elements_category is None:
            elements_category         = []

        if weight_category is None:
            weight_category           = []

        if elements_problem_category is None:
            elements_problem_category = []

        if weight_problem_category is None:
            weight_problem_category   = []

        if elements_affected_devices is None:
            elements_affected_devices = []

        if weight_affected_devices is None:
            weight_affected_devices = []

        if elements_reason_for_creating is None:
            elements_reason_for_creating = []

        if weights_reason_for_creating is None:
            weights_reason_for_creating = []

        if elements_issue_status is None:
            elements_issue_status = []

        if weights_issue_status is None:
            weights_issue_status = []

        if elements_sla_met is None:
            elements_sla_met = []

        if weights_sla_met is None:
            weights_sla_met = []

        if elements_priority is None:
            elements_priority = []

        if weights_priority is None:
            weights_priority = []

        # ------------------------- Other methods ---------------------------------#

        def print_full_name(self):
            return self.first_name+" "+self.last_name

        #def weights_category(self):


def main():
    print("Preparing workbook...")
    wb = load_workbook(
        r"C:\Users\abautista\PycharmProjects\apps\backbone_navigator_github\Data_Visualization_Framework\Fake data\navigator_data_daily_insertion.xlsx")
    print("Workbook is ready!")
    incident_counter = determine_incident_counter(wb)
    engineer_one, engineer_two, engineer_three, engineer_four, engineer_five = generate_engineers_data(incident_counter)
    insert_data_in_worksheet(engineer_one, engineer_two, engineer_three, engineer_four, engineer_five, wb)

def generate_engineers_data(incident_counter):
    ''' The following are a brief description of the engineers that will be simulated based on their
            probability distributions.

        Name: Alekz
        Last Name: Horne
        Tickets solved per day: 7 - 9
        Weeks that were considered into account: week 29 - week 52 from 2017

        Category: IoT
            Weight(s): 1

        Problem category: Breach in SCADA systems,Compromised data,Erroneous data load,
                          Hijack of device, Internal flaws in device, Malware installed,
                          Penetration breach
            Weight(s): [0.065, 0.232, 0.094, 0.082, 0.034, 0.165,0.328]

        Affected devices: Company main devices, Personal user devices, Third party devices
            Weight(s): [0.117, 0.775, 0.108]

        Reason for creating: Daily Analysis, Hourly Analysis, Special Request
            Weight(s): [0.754, 0.092, 0.154]

        Issue_status : Closed, In Progress, Transferred
            Weight(s): [0.117, 0.093, 0.790]

        SLA_MET: Yes
            Weight(s): 1

        Priority = "P1", "P2", "P3"
            Weight(s): [0.082, 0.293, 0.625]

        KPI_I: Range goes from 0 - 45
            Wieght(s): They were not calculated

        KPI_II: Range goes from 0 - 24
            Wieght(s): They were not calculated

        KPI_III: Range goes from 0 - 31
            Wieght(s): They were not calculated

        KPI_IV: Range goes from 0 - 28
            Wieght(s): They were not calculated


        Name: Annette
        Last Name: Smith
        Tickets solved per day: 6 - 8
        Weeks that were considered into account: week 25 - week 52 from 2017

        Category: Cybersecurity
            Weight(s): 1

        Problem category: "Bad security policies", "Compromised data", "DDoS attack", "Erroneous data load",
                          "Identity spoofing", "Lack of CERT", "Penetration breach", "Ransomware virus"
            Weight(s): [0.267, 0.017, 0.156, 0.004, 0.003, 0.037, 0.146, 0.371]


        Affected devices: "Company main devices", "Personal user devices", "Third party devices"
            Weight(s): [0.326, 0.530, 0.145]

        Reason for creating: ["Bi-weekly Analysis", "Daily Analysis", "Hourly Analysis", "Random Analysis", "Special Request"]
            Weight(s): [0.001, 0.112, 0.644, 0.121, 0.122]

        Issue_status : "Closed", "In Progress", "Transferred"]
            Weight(s): [0.327, 0.011, 0.662]

        SLA_MET: Yes
            Weight(s): 1

        Priority = "P2", "P3"
            Weight(s): [0.334, 0.666]

        KPI_I: Range goes from 0 - 45
            Wieght(s): They were not calculated

        KPI_II: Range goes from 0 - 48
            Wieght(s): They were not calculated

        KPI_III: Range goes from 0 - 64
            Wieght(s): They were not calculated

        KPI_IV: Range goes from 0 - 47
            Wieght(s): They were not calculated

        Name: Dieter
        Last Name: Becker
        Tickets solved per day: 8 - 10
        Weeks that were considered into account: week 23 - week 52 from 2017

        Category: IoT
            Weight(s): 1

        Problem category: ["Breach in SCADA systems", "Compromised data", "Erroneous data load", "Hijack of device"
                         "Internal flaws in device", "Malware installed", "Penetration breach"]
            Weight(s): [0.019, 0.315, 0.011, 0.060, 0.002, 0.079, 0.514]

        Affected devices: ["Company main devices", "Mobile devices", "Personal user devices", "Principal devices"
                          "Third party devices"]
            Weight(s): [0.158, 0.001, 0.826, 0.001, 0.015]

        Reason for creating: ["Daily Analysis", "Hourly Analysis", "Immediate Analysis", "Special Request", "VIP"]
            Weight(s): [0.623, 0.221, 0.011, 0.141, 0.005]

        Issue_status : "Closed", "In Progress", "Transferred"]
            Weight(s): [0.151, 0.046, 0.803]

        SLA_MET: Yes
            Weight(s): 1

        Priority = "P1", "P2", "P3"
            Weight(s): [0.119, 0.789, 0.092]

        KPI_I: Range goes from 0 - 47
            Wieght(s): They were not calculated

        KPI_II: Range goes from 0 - 73
            Wieght(s): They were not calculated

        KPI_III: Range goes from 0 - 72
            Wieght(s): They were not calculated

        KPI_IV: Range goes from 0 - 114
            Wieght(s): They were not calculated

        Name: Florian
        Last Name: Klein
        Tickets solved per day: 8 - 10
        Weeks that were considered into account: week 32 - week 52 from 2017. Week 42 is not considered.

        Category: IoT
            Weight(s): 1

        Problem category: ["Breach in SCADA systems", "Compromised data", "Erroneous data load", "Hijack of device",
                            "Internal flaws in device", "Malware installed", "Penetration breach"]
            Weight(s): [0.016, 0.186, 0.004, 0.117, 0.004, 0.277, 0.396]

        Affected devices: ["Company main devices", "Mobile devices", "Personal user devices", "Third party devices"
                          "Third party devices"]
            Weight(s): [0.333, 0.004, 0.607, 0.056]

        Reason for creating: ["Daily Analysis", "Hourly Analysis", "Special Request"]
            Weight(s): [0.807, 0.128, 0.065]

        Issue_status : ["Closed", "In Progress", "Transferred"]
            Weight(s): [0.333, 0.002, 0.665]

        SLA_MET: Yes
            Weight(s): 1

        Priority = "P1", "P2", "P3"
            Weight(s): [0.007, 0.077, 0.915]

        KPI_I: Range goes from 0 - 37
            Wieght(s): They were not calculated

        KPI_II: Range goes from 0 - 2
            Wieght(s): They were not calculated

        KPI_III: Range goes from 0 - 72
            Wieght(s): They were not calculated

        KPI_IV: Range goes from 0 - 67
            Wieght(s): They were not calculated

        Name: Hannes
        Last Name: Weber
        Tickets solved per day: 6 - 7
        Weeks that were considered into account: week 1 and week 32- week 52 from 2017. Week 42 is not considered.

        Category: IoT
            Weight(s): 1

        Problem category: ["Breach in SCADA systems", "Compromised data", "Erroneous data load", "Hijack of device",
                           "Internal flaws in device",  "Malware installed",  "Penetration breach"]
            Weight(s): [0.026, 0.313, 0.017, 0.012, 0.017, 0.388, 0.228]

        Affected devices: ["Company main devices",  "Personal user devices",  "Third party devices"]
            Weight(s): [0.173, 0.774, 0.053]

        Reason for creating: ["Daily Analysis", "Hourly Analysis", "Immediate Analysis", "Special Request", "VIP"]
            Weight(s): [0.743, 0.002, 0.179, 0.077]

        Issue_status : ["Closed", "In Progress", "Open", "Transferred"]
            Weight(s): [0.170, 0.014, 0.003, 0.813]

        SLA_MET: ["Yes", "No"]
            Weight(s): [0.995, 0.005]

        Priority = ["P0", "P1", "P2", "P3"]
            Weight(s): [0.002, 0.071, 0.165, 0.762]

        KPI_I: Range goes from 0 - 37
            Wieght(s): They were not calculated

        KPI_II: Range goes from 0 - 25
            Wieght(s): They were not calculated

        KPI_III: Range goes from 0 - 105
            Wieght(s): They were not calculated

        KPI_IV: Range goes from 0 - 66
            Wieght(s): They were not calculated

    '''
    # ---------------------- Ask for the initial week where you want to display the tickets -----------------#

    week = int(input("Please, provide the week number where you want to insert the new tickets: "))
    year = int(input("Please, provide the year where you want to insert the new tickets: "))
    initial_date, ending_date = get_start_end_date(year, week)
    converted_initial_date, converted_ending_date = convert_date_to_datetime(initial_date, ending_date)


    # ------------------------------- Initialize main components of each engineer --------------------------#

    '''
        The lists of constructors lists are not declared here but the number of tickets solved per day are declared.
    '''
    engineer_alekz_horne   = Engineer("Alekz","Horne", random.randint(7,9))
    engineer_annette_smith = Engineer("Annette","Smith", random.randint(6,8))
    engineer_hannes_weber  = Engineer("Hannes", "Weber", random.randint(6,7))
    engineer_florian_klein = Engineer("Florian", "Klein", random.randint(8, 10))
    engineer_dieter_becker = Engineer("Dieter", "Becker", random.randint(8, 10))


    generated_tickets_alekz_horne   = []
    generated_tickets_annette_smith = []
    generated_tickets_hannes_weber  = []
    generated_tickets_florian_klein = []
    generated_tickets_dieter_becker = []

    #---------------------------------Generate simulated tickets for Alekz Horne-------------------------------------#


    print(engineer_alekz_horne.first_name)
    print(engineer_alekz_horne.last_name)
    print(engineer_alekz_horne.amount_of_tickets_solved_day)

    engineer_alekz_horne.elements_category = ["IoT"]
    engineer_alekz_horne.weight_category = [1]

    engineer_alekz_horne.elements_problem_category = ["Breach in SCADA systems","Compromised data","Erroneous data load",
                                                      "Hijack of device", "Internal flaws in device", "Malware installed",
                                                      "Penetration breach"]
    engineer_alekz_horne.weights_problem_category  = [0.065, 0.232, 0.094, 0.082, 0.034, 0.165,0.328]

    engineer_alekz_horne.elements_affected_devices = ["Company main devices", "Personal user devices", "Third party devices"]
    engineer_alekz_horne.weights_affected_devices = [0.117, 0.775, 0.108]

    engineer_alekz_horne.elements_reason_for_creating = ["Daily Analysis", "Hourly Analysis","Special Request"]
    engineer_alekz_horne.weights_reason_for_creating = [0.754, 0.092, 0.154]


    engineer_alekz_horne.elements_issue_status = ["Closed","In Progress","Transferred"]
    engineer_alekz_horne.weights_issue_status  = [0.117, 0.093, 0.790]

    engineer_alekz_horne.elements_sla_met = ["Yes"]
    engineer_alekz_horne.weights_sla_met  = [1]

    engineer_alekz_horne.elements_priority = ["P1","P2","P3"]
    engineer_alekz_horne.weights_priority = [0.082, 0.293, 0.625]

    for day in range(int(engineer_alekz_horne.amount_of_tickets_solved_day)):
        incident_counter = incident_counter + 1
        generated_tickets_alekz_horne.append(incident_counter)
        generated_tickets_alekz_horne.append(engineer_alekz_horne.first_name + " " + engineer_alekz_horne.last_name)
        generated_tickets_alekz_horne.append("F-Secure")
        generated_tickets_alekz_horne.append(choice(engineer_alekz_horne.elements_category, p=engineer_alekz_horne.weight_category))
        generated_tickets_alekz_horne.append(
            choice(engineer_alekz_horne.elements_problem_category, p=engineer_alekz_horne.weights_problem_category))

        # issue status
        issue_status = choice(engineer_alekz_horne.elements_issue_status, p=engineer_alekz_horne.weights_issue_status)
        generated_tickets_alekz_horne.append(issue_status)

        # Initial Detection Date
        initial_detection_date = create_random_date(converted_initial_date, converted_ending_date)
        generated_tickets_alekz_horne.append(initial_detection_date)

        # Initial Action Date - You need to create an action date after the incident was first initiated
        action_date = create_random_date(initial_detection_date, converted_ending_date)
        generated_tickets_alekz_horne.append(action_date)

        # Issue Closed Date or Issue Re-Assigned Date
        if issue_status == "Closed" or issue_status == "Transferred":
            generated_tickets_alekz_horne.append(create_random_date(action_date, converted_ending_date))
        # Open or Queued
        else:
            generated_tickets_alekz_horne.append("NaN")


        generated_tickets_alekz_horne.append(choice(engineer_alekz_horne.elements_affected_devices, p=engineer_alekz_horne.weights_affected_devices))
        generated_tickets_alekz_horne.append(choice(engineer_alekz_horne.elements_reason_for_creating, p=engineer_alekz_horne.weights_reason_for_creating))
        #generated_tickets_alekz_horne.append(choice(engineer_alekz_horne.elements_issue_status, p=engineer_alekz_horne.weights_issue_status))
        generated_tickets_alekz_horne.append(choice(engineer_alekz_horne.elements_sla_met, p=engineer_alekz_horne.weights_sla_met))
        generated_tickets_alekz_horne.append(choice(engineer_alekz_horne.elements_priority, p=engineer_alekz_horne.weights_priority))
        #generated_tickets_alekz_horne.append("KPI_I")
        generated_tickets_alekz_horne.append(random.randint(0, 45))
        #generated_tickets_alekz_horne.append("KPI_II")
        generated_tickets_alekz_horne.append(random.randint(0, 24))
        #generated_tickets_alekz_horne.append("KPI_III")
        generated_tickets_alekz_horne.append(random.randint(0, 31))
        #generated_tickets_alekz_horne.append("KPI_IV")
        generated_tickets_alekz_horne.append(random.randint(0, 28))
        generated_tickets_alekz_horne.append(";")
        #print(choice(engineer_alekz_horne.elements_category, p=engineer_alekz_horne.weight_category))
    print(generated_tickets_alekz_horne)

    #---------------------------------Generate tickets for Annette Smith -------------------------------------#

    print(engineer_annette_smith.first_name)
    print(engineer_annette_smith.last_name)
    print(engineer_annette_smith.amount_of_tickets_solved_day)

    engineer_annette_smith.elements_category = ["CyberSecurity"]
    engineer_annette_smith.weight_category = [1]

    engineer_annette_smith.elements_problem_category = ["Bad security policies", "Compromised data",
                                                      "DDoS attack","Erroneous data load",
                                                      "Identity spoofing","Lack of CERT",
                                                      "Penetration breach", "Ransomware virus"]

    engineer_annette_smith.weights_problem_category = [0.2667, 0.0167, 0.1555, 0.0039, 0.0030, 0.0374, 0.1457, 0.3711]

    engineer_annette_smith.elements_affected_devices = ["Company main devices", "Personal user devices",
                                                      "Third party devices"]
    engineer_annette_smith.weights_affected_devices = [0.3258, 0.5295, 0.1447]

    engineer_annette_smith.elements_reason_for_creating = ["Bi-weekly Analysis","Daily Analysis", "Hourly Analysis", "Random Analysis", "Special Request"]
    engineer_annette_smith.weights_reason_for_creating = [0.001, 0.112, 0.644, 0.121, 0.122]

    engineer_annette_smith.elements_issue_status = ["Closed", "In Progress", "Transferred"]
    engineer_annette_smith.weights_issue_status = [0.327, 0.011, 0.662]

    engineer_annette_smith.elements_sla_met = ["Yes"]
    engineer_annette_smith.weights_sla_met = [1]

    engineer_annette_smith.elements_priority = ["P2", "P3"]
    engineer_annette_smith.weights_priority = [0.3337,0.6663]

    for day in range(int(engineer_annette_smith.amount_of_tickets_solved_day)):
        incident_counter = incident_counter + 1
        generated_tickets_annette_smith.append(incident_counter)
        generated_tickets_annette_smith.append(engineer_annette_smith.first_name + " " +engineer_annette_smith.last_name)
        generated_tickets_annette_smith.append("F-Secure")
        generated_tickets_annette_smith.append(
            choice(engineer_annette_smith.elements_category, p=engineer_annette_smith.weight_category))
        generated_tickets_annette_smith.append(
            choice(engineer_annette_smith.elements_problem_category, p=engineer_annette_smith.weights_problem_category))

        # issue status
        issue_status = choice(engineer_annette_smith.elements_issue_status,
                              p=engineer_annette_smith.weights_issue_status)
        generated_tickets_annette_smith.append(issue_status)

        # Initial Detection Date
        initial_detection_date = create_random_date(converted_initial_date, converted_ending_date)
        generated_tickets_annette_smith.append(initial_detection_date)

        # Initial Action Date - You need to create an action date after the incident was first initiated
        action_date = create_random_date(initial_detection_date, converted_ending_date)
        generated_tickets_annette_smith.append(action_date)

        # Issue Closed Date or Issue Re-Assigned Date
        if issue_status == "Closed" or issue_status == "Transferred":
            generated_tickets_annette_smith.append(create_random_date(action_date, converted_ending_date))
        # Open or Queued
        else:
            generated_tickets_annette_smith.append("NaN")

        generated_tickets_annette_smith.append(
            choice(engineer_annette_smith.elements_affected_devices, p=engineer_annette_smith.weights_affected_devices))
        generated_tickets_annette_smith.append(choice(engineer_annette_smith.elements_reason_for_creating,
                                                    p=engineer_annette_smith.weights_reason_for_creating))
        #generated_tickets_annette_smith.append(choice(engineer_annette_smith.elements_issue_status, p=engineer_annette_smith.weights_issue_status))
        generated_tickets_annette_smith.append(
            choice(engineer_annette_smith.elements_sla_met, p=engineer_annette_smith.weights_sla_met))
        generated_tickets_annette_smith.append(
            choice(engineer_annette_smith.elements_priority, p=engineer_annette_smith.weights_priority))
        #generated_tickets_annette_smith.append("KPI_I")
        generated_tickets_annette_smith.append(random.randint(0, 45)) # KPIs should be adjusted to random choice to get more accurate data
        #generated_tickets_annette_smith.append("KPI_II")
        generated_tickets_annette_smith.append(random.randint(0, 48))
        #generated_tickets_annette_smith.append("KPI_III")
        generated_tickets_annette_smith.append(random.randint(0, 64))
        #generated_tickets_annette_smith.append("KPI_IV")
        generated_tickets_annette_smith.append(random.randint(0, 47))
        generated_tickets_annette_smith.append(";")
        # print(choice(engineer_alekz_horne.elements_category, p=engineer_alekz_horne.weight_category))
    print(generated_tickets_annette_smith)


    #---------------------------------Generate tickets for Florian Klein-------------------------------------#

    print(engineer_florian_klein.first_name)
    print(engineer_florian_klein.last_name)
    print(engineer_florian_klein.amount_of_tickets_solved_day)

    engineer_florian_klein.elements_category = ["CyberSecurity"]
    engineer_florian_klein.weight_category = [1]

    engineer_florian_klein.elements_problem_category = ["Breach in SCADA systems", "Compromised data", "Erroneous data load", "Hijack of device",
                            "Internal flaws in device", "Malware installed", "Penetration breach"]

    engineer_florian_klein.weights_problem_category = [0.016, 0.186, 0.004, 0.117, 0.004, 0.277, 0.396]

    engineer_florian_klein.elements_affected_devices = ["Company main devices", "Mobile devices", "Personal user devices", "Third party devices"]
    engineer_florian_klein.weights_affected_devices = [0.333, 0.004, 0.607, 0.056]

    engineer_florian_klein.elements_reason_for_creating = ["Daily Analysis", "Hourly Analysis", "Special Request"]
    engineer_florian_klein.weights_reason_for_creating = [0.807, 0.128, 0.065]

    engineer_florian_klein.elements_issue_status = ["Closed", "In Progress", "Transferred"]
    engineer_florian_klein.weights_issue_status = [0.333, 0.002, 0.665]

    engineer_florian_klein.elements_sla_met = ["Yes"]
    engineer_florian_klein.weights_sla_met = [1]

    engineer_florian_klein.elements_priority = ["P1", "P2", "P3"]
    engineer_florian_klein.weights_priority = [0.0072, 0.0775, 0.9153]

    for day in range(int(engineer_florian_klein.amount_of_tickets_solved_day)):
        incident_counter = incident_counter + 1
        generated_tickets_florian_klein.append(incident_counter)
        generated_tickets_florian_klein.append(engineer_florian_klein.first_name + " " + engineer_florian_klein.last_name)
        generated_tickets_florian_klein.append("F-Secure")
        generated_tickets_florian_klein.append(
            choice(engineer_florian_klein.elements_category, p=engineer_florian_klein.weight_category))
        generated_tickets_florian_klein.append(
            choice(engineer_florian_klein.elements_problem_category, p=engineer_florian_klein.weights_problem_category))

        # issue status
        issue_status = choice(engineer_florian_klein.elements_issue_status, p=engineer_florian_klein.weights_issue_status)
        generated_tickets_florian_klein.append(issue_status)

        # Initial Detection Date
        initial_detection_date = create_random_date(converted_initial_date, converted_ending_date)
        generated_tickets_florian_klein.append(initial_detection_date)

        # Initial Action Date - You need to create an action date after the incident was first initiated
        action_date = create_random_date(initial_detection_date, converted_ending_date)
        generated_tickets_florian_klein.append(action_date)

        # Issue Closed Date or Issue Re-Assigned Date
        if issue_status == "Closed" or issue_status == "Transferred":
            generated_tickets_florian_klein.append(create_random_date(action_date, converted_ending_date))
        # Open or Queued
        else:
            generated_tickets_florian_klein.append("NaN")

        generated_tickets_florian_klein.append(
            choice(engineer_florian_klein.elements_affected_devices, p=engineer_florian_klein.weights_affected_devices))
        generated_tickets_florian_klein.append(choice(engineer_florian_klein.elements_reason_for_creating,
                                                    p=engineer_florian_klein.weights_reason_for_creating))
        #generated_tickets_florian_klein.append(choice(engineer_florian_klein.elements_issue_status, p=engineer_florian_klein.weights_issue_status))
        generated_tickets_florian_klein.append(
            choice(engineer_florian_klein.elements_sla_met, p=engineer_florian_klein.weights_sla_met))
        generated_tickets_florian_klein.append(
            choice(engineer_florian_klein.elements_priority, p=engineer_florian_klein.weights_priority))
        #generated_tickets_florian_klein.append("KPI_I")
        generated_tickets_florian_klein.append(random.randint(0, 37)) # KPIs should be adjusted to random choice to get more accurate data
        #generated_tickets_florian_klein.append("KPI_II")
        generated_tickets_florian_klein.append(random.randint(0, 2))
        #generated_tickets_florian_klein.append("KPI_III")
        generated_tickets_florian_klein.append(random.randint(0, 72))
        #generated_tickets_florian_klein.append("KPI_IV")
        generated_tickets_florian_klein.append(random.randint(0, 67))
        generated_tickets_florian_klein.append(";")
        # print(choice(engineer_alekz_horne.elements_category, p=engineer_alekz_horne.weight_category))
    print(generated_tickets_florian_klein)

    # ---------------------------------Generate tickets for Dieter Becker -------------------------------------#
    '''
    print(engineer_dieter_becker.first_name)
    print(engineer_dieter_becker.last_name)
    print(engineer_dieter_becker.amount_of_tickets_solved_day)

    engineer_dieter_becker.elements_category = ["IoT"]
    engineer_dieter_becker.weight_category = [1]

    engineer_dieter_becker.elements_problem_category = ["Breach in SCADA systems", "Compromised data", "Erroneous data load", "Hijack of device",
                         "Internal flaws in device", "Malware installed", "Penetration breach"]

    engineer_dieter_becker.weights_problem_category = [0.019, 0.315, 0.011, 0.060, 0.002, 0.079, 0.514]

    engineer_dieter_becker.elements_affected_devices = ["Company main devices", "Mobile devices", "Personal user devices", "Principal devices",
                                                        "Third party devices"]
    engineer_dieter_becker.weights_affected_devices = [0.1576, 0.0010, 0.8259, 0.0010, 0.0145]

    engineer_dieter_becker.elements_reason_for_creating = ["Daily Analysis", "Hourly Analysis", "Immediate Analysis", "Special Request", "VIP"]
    engineer_dieter_becker.weights_reason_for_creating = [0.6228, 0.2205, 0.0106, 0.1412, 0.0048]
    #print(np.sum([0.6228, 0.2205, 0.0106, 0.1412, 0.0048]))

    engineer_dieter_becker.elements_issue_status = ["Closed", "In Progress", "Transferred"]
    engineer_dieter_becker.weights_issue_status = [0.1509, 0.0464, 0.8027]

    engineer_dieter_becker.elements_sla_met = ["Yes"]
    engineer_dieter_becker.weights_sla_met = [1]

    engineer_dieter_becker.elements_priority = ["P1", "P2", "P3"]
    engineer_dieter_becker.weights_priority = [0.119, 0.789, 0.092]

    for day in range(int(engineer_dieter_becker.amount_of_tickets_solved_day)):
        generated_tickets_dieter_becker.append("F-Secure")
        generated_tickets_dieter_becker.append(
            choice(engineer_dieter_becker.elements_category, p=engineer_dieter_becker.weight_category))
        generated_tickets_dieter_becker.append(
            choice(engineer_dieter_becker.elements_problem_category, p=engineer_dieter_becker.weights_problem_category))
        generated_tickets_dieter_becker.append(create_random_date(initial_date, ending_date))
        generated_tickets_dieter_becker.append(
            choice(engineer_dieter_becker.elements_affected_devices, p=engineer_dieter_becker.weights_affected_devices))
        generated_tickets_dieter_becker.append(choice(engineer_dieter_becker.elements_reason_for_creating,
                                                    p=engineer_dieter_becker.weights_reason_for_creating))
        generated_tickets_dieter_becker.append(
            choice(engineer_dieter_becker.elements_issue_status, p=engineer_dieter_becker.weights_issue_status))
        generated_tickets_dieter_becker.append(
            choice(engineer_dieter_becker.elements_sla_met, p=engineer_dieter_becker.weights_sla_met))
        generated_tickets_dieter_becker.append(
            choice(engineer_dieter_becker.elements_priority, p=engineer_dieter_becker.weights_priority))
        generated_tickets_dieter_becker.append("KPI_I")
        generated_tickets_dieter_becker.append(random.randint(0, 47)) # KPIs should be adjusted to random choice to get more accurate data
        generated_tickets_dieter_becker.append("KPI_II")
        generated_tickets_dieter_becker.append(random.randint(0, 73))
        generated_tickets_dieter_becker.append("KPI_III")
        generated_tickets_dieter_becker.append(random.randint(0, 72))
        generated_tickets_dieter_becker.append("KPI_IV")
        generated_tickets_dieter_becker.append(random.randint(0, 114))
        generated_tickets_dieter_becker.append(";")
        # print(choice(engineer_alekz_horne.elements_category, p=engineer_alekz_horne.weight_category))
    print(generated_tickets_dieter_becker)
    '''

    # ---------------------------------Generate tickets for Hannes Weber -------------------------------------#

    print(engineer_hannes_weber.first_name)
    print(engineer_hannes_weber.last_name)
    print(engineer_hannes_weber.amount_of_tickets_solved_day)

    engineer_hannes_weber.elements_category = ["IoT"]
    engineer_hannes_weber.weight_category = [1]


    engineer_hannes_weber.elements_problem_category = ["Breach in SCADA systems", "Compromised data", "Erroneous data load", "Hijack of device",
                                                       "Internal flaws in device", "Malware installed", "Penetration breach"]

    engineer_hannes_weber.weights_problem_category = [0.0255, 0.3129, 0.0170, 0.0119, 0.0170, 0.3878, 0.2279]

    engineer_hannes_weber.elements_affected_devices = ["Company main devices",  "Personal user devices", "Third party devices"]
    engineer_hannes_weber.weights_affected_devices = [0.1735, 0.7738, 0.0527]

    engineer_hannes_weber.elements_reason_for_creating = ["Daily Analysis", "Halt Analysis", "Hourly Analysis", "Special Request"]
    engineer_hannes_weber.weights_reason_for_creating = [0.7432, 0.0017, 0.1786, 0.0765]

    engineer_hannes_weber.elements_issue_status = ["Closed", "In Progress", "Open", "Transferred"]
    engineer_hannes_weber.weights_issue_status = [0.170, 0.014, 0.003, 0.813]

    engineer_hannes_weber.elements_sla_met = ["Yes", "No"]
    engineer_hannes_weber.weights_sla_met = [0.995, 0.005]

    engineer_hannes_weber.elements_priority = ["P0", "P1", "P2", "P3"]
    engineer_hannes_weber.weights_priority = [0.002, 0.071, 0.165, 0.762]

    for day in range(int(engineer_hannes_weber.amount_of_tickets_solved_day)):
        incident_counter = incident_counter + 1
        generated_tickets_hannes_weber.append(incident_counter)
        generated_tickets_hannes_weber.append(engineer_hannes_weber.first_name +" " + engineer_hannes_weber.last_name)
        generated_tickets_hannes_weber.append("F-Secure")
        generated_tickets_hannes_weber.append(
            choice(engineer_hannes_weber.elements_category, p=engineer_hannes_weber.weight_category))
        generated_tickets_hannes_weber.append(
            choice(engineer_hannes_weber.elements_problem_category, p=engineer_hannes_weber.weights_problem_category))
        #issue status
        issue_status = choice(engineer_hannes_weber.elements_issue_status, p=engineer_hannes_weber.weights_issue_status)
        generated_tickets_hannes_weber.append(issue_status)

        # Initial Detection Date
        initial_detection_date = create_random_date(converted_initial_date, converted_ending_date)
        generated_tickets_hannes_weber.append(initial_detection_date)

        # Initial Action Date - You need to create an action date after the incident was first initiated
        action_date = create_random_date(initial_detection_date, converted_ending_date)
        generated_tickets_hannes_weber.append(action_date)

        # Issue Closed Date or Issue Re-Assigned Date
        if issue_status == "Closed" or issue_status == "Transferred":
            generated_tickets_hannes_weber.append(create_random_date(action_date, converted_ending_date))
        # Open or Queued
        else:
            generated_tickets_hannes_weber.append("NaN")

        generated_tickets_hannes_weber.append(
            choice(engineer_hannes_weber.elements_affected_devices, p=engineer_hannes_weber.weights_affected_devices))

        generated_tickets_hannes_weber.append(choice(engineer_hannes_weber.elements_reason_for_creating,
                                                    p=engineer_hannes_weber.weights_reason_for_creating))
        generated_tickets_hannes_weber.append(
            choice(engineer_hannes_weber.elements_sla_met, p=engineer_hannes_weber.weights_sla_met))
        generated_tickets_hannes_weber.append(
            choice(engineer_hannes_weber.elements_priority, p=engineer_hannes_weber.weights_priority))
        #generated_tickets_hannes_weber.append("KPI_I")
        generated_tickets_hannes_weber.append(random.randint(0, 47)) # KPIs should be adjusted to random choice to get more accurate data
        #generated_tickets_hannes_weber.append("KPI_II")
        generated_tickets_hannes_weber.append(random.randint(0, 73))
        #generated_tickets_hannes_weber.append("KPI_III")
        generated_tickets_hannes_weber.append(random.randint(0, 72))
        #generated_tickets_hannes_weber.append("KPI_IV")
        generated_tickets_hannes_weber.append(random.randint(0, 114))
        generated_tickets_hannes_weber.append(";")
    print(generated_tickets_hannes_weber)

    return generated_tickets_alekz_horne, generated_tickets_annette_smith, generated_tickets_florian_klein, \
           generated_tickets_hannes_weber, generated_tickets_dieter_becker


def insert_data_in_worksheet(engineer_one, engineer_two, engineer_three, engineer_four, engineer_five, wb):

    #print("Loading workbook, please wait...")
    # load the existing spreadsheet - no need to add double \\
    #wb = load_workbook(r"C:\Users\abautista\PycharmProjects\apps\backbone_navigator_github\Data_Visualization_Framework\Fake data\navigator_data_daily_insertion.xlsx")
    #print("The workbook has been loaded.")

    #get the current worksheet
    write_current_sheet = wb.active

    #define the last value of rows
    row_count = write_current_sheet.max_row

    # Take column 1 as a reference to get all the rows
    #for index, row in enumerate(write_current_sheet.iter_rows()):
    #    for cell in row:
    #        print(write_current_sheet.cell(row=index+1, column=1).value, cell.value)

    # declare this variable outside of the loop to avoid any data corruption
    last_value_column_one = 0
    # iterate ONLY over column 1 which contains the incident IDs
    for i in range(row_count):
        #print(write_current_sheet.cell(row=i+1, column=1).value)
        last_value_column_one = write_current_sheet.cell(row=i+1, column=1).value
    print(last_value_column_one)

def determine_incident_counter(wb):
    # get the current worksheet
    write_current_sheet = wb.active

    # define the last value of rows
    row_count = write_current_sheet.max_row

    last_value_column_one = 0
    for i in range(row_count):
        last_value_column_one = write_current_sheet.cell(row=i+1, column=1).value

    return last_value_column_one

def get_start_end_date(year, week):
    '''Given the year and week, return the first and last day of the given week and year.
       The first day is Sunday and last day is Saturday'''

    year = int(year)
    week = int(week)

    d = date(year, 1, 1)
    # patch to fix the problem of the retrieval of data for the first weeks of the year
    dlt = timedelta(days=(week - 1) * 7)

    if (d.weekday() <= 3):
        d = d - timedelta(d.weekday())
    else:
        # this line  d = d + timedelta(6 - d.weekday()) indicates to start in Sunday and end in Saturday
        # but it breaks when the retrieval is during the first weeks of the year, so we have to change the code
        # back to d = d + timedelta(7 - d.weekday())
        d = d + timedelta(7 - d.weekday())
    return d + dlt, d + dlt + timedelta(days=6)

def create_random_date(start, end):
    """
        This function will return a random datetime between two datetime objects.
    """
    delta = end - start
    int_delta = (delta.days * 24 * 60 * 60) + delta.seconds
    random_second = randrange(int_delta)
    return start + timedelta(seconds=random_second)

    #return datetime.strptime(str(start + timedelta(seconds=random_second)), '%Y-%m-%d')

def convert_date_to_datetime(beginning_date,ending_date):
    required_space = " "
    beginning_time_string_format = "00:00:00"
    ending_time_string_format    = "23:59:59"
    beginning_date_date_format_to_string = str(beginning_date).split(" ")[0] + required_space + beginning_time_string_format
    ending_date_date_format_to_string = str(ending_date).split(" ")[0] + str() + required_space + ending_time_string_format

    ####################### convert the string to datetime.datetime ##################
    beginning_date_string_to_datetime = datetime.strptime(beginning_date_date_format_to_string, "%Y-%m-%d %H:%M:%S")
    ending_date_string_to_datetime = datetime.strptime(ending_date_date_format_to_string, "%Y-%m-%d %H:%M:%S")

    return beginning_date_string_to_datetime, ending_date_string_to_datetime

if __name__ == "__main__":
    main()
